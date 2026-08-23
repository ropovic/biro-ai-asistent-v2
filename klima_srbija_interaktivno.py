import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==============================================================================
# BAZA GLAVNIH METEOROLOŠKIH STANICA U SRBIJI (GMS RHMZ)
# ==============================================================================
STANICE = {
    1:  {"naziv": "BEOGRAD - VRAČAR", "lat": 44.79, "lon": 20.46, "alt": 132},
    2:  {"naziv": "NOVI SAD - RIMSKI ŠANČEVI", "lat": 45.33, "lon": 19.85, "alt": 84},
    3:  {"naziv": "NIŠ", "lat": 43.33, "lon": 21.90, "alt": 202},
    4:  {"naziv": "KRAGUJEVAC", "lat": 44.02, "lon": 20.92, "alt": 185},
    5:  {"naziv": "SUBOTICA", "lat": 46.10, "lon": 19.67, "alt": 108},
    6:  {"naziv": "SOMBOR", "lat": 45.77, "lon": 19.12, "alt": 87},
    7:  {"naziv": "ZRENJANIN", "lat": 45.38, "lon": 20.40, "alt": 80},
    8:  {"naziv": "KIKINDA", "lat": 45.82, "lon": 20.47, "alt": 81},
    9:  {"naziv": "BANATSKI KARLOVAC", "lat": 45.05, "lon": 21.03, "alt": 96},
    10: {"naziv": "VRŠAC", "lat": 45.12, "lon": 21.30, "alt": 84},
    11: {"naziv": "SREMSKA MITROVICA", "lat": 44.97, "lon": 19.61, "alt": 82},
    12: {"naziv": "VALJEVO", "lat": 44.27, "lon": 19.88, "alt": 176},
    13: {"naziv": "LOZNICA", "lat": 44.53, "lon": 19.23, "alt": 121},
    14: {"naziv": "SMEDEREVSKA PALANKA", "lat": 44.37, "lon": 20.95, "alt": 121},
    15: {"naziv": "VELIKO GRADIŠTE", "lat": 44.75, "lon": 21.52, "alt": 82},
    16: {"naziv": "CRNI VRH", "lat": 44.12, "lon": 21.96, "alt": 1027},
    17: {"naziv": "NEGOTIN", "lat": 44.23, "lon": 22.53, "alt": 42},
    18: {"naziv": "ZLATIBOR", "lat": 43.65, "lon": 19.70, "alt": 1028},
    19: {"naziv": "SJENICA", "lat": 43.27, "lon": 19.99, "alt": 1038},
    20: {"naziv": "POŽEGA", "lat": 43.85, "lon": 20.03, "alt": 310},
    21: {"naziv": "KRALJEVO", "lat": 43.72, "lon": 20.68, "alt": 215},
    22: {"naziv": "KOPAONIK", "lat": 43.28, "lon": 20.82, "alt": 1710},
    23: {"naziv": "KURŠUMLIJA", "lat": 43.13, "lon": 21.27, "alt": 384},
    24: {"naziv": "KRUŠEVAC", "lat": 43.58, "lon": 21.32, "alt": 166},
    25: {"naziv": "ĆUPRIJA", "lat": 43.93, "lon": 21.37, "alt": 123},
    26: {"naziv": "LESKOVAC", "lat": 43.00, "lon": 21.95, "alt": 230},
    27: {"naziv": "ZAJEČAR", "lat": 43.88, "lon": 22.28, "alt": 144},
    28: {"naziv": "DIMITROVGRAD", "lat": 43.02, "lon": 22.75, "alt": 450},
    29: {"naziv": "VRANJE", "lat": 42.55, "lon": 21.90, "alt": 432},
    30: {"naziv": "PRIŠTINA", "lat": 42.65, "lon": 21.15, "alt": 573}
}

# ==============================================================================
# INTERAKTIVNI IZBOR KORISNIKA
# ==============================================================================
def izaberi_stanicu_i_period():
    print("\n" + "="*60)
    print(" KLIMATOLOŠKI PODACI ZA METEOROLOŠKE STANICE U SRBIJI")
    print("="*60)
    
    for key, val in STANICE.items():
        print(f"[{key:2d}] {val['naziv']:<26} (lat: {val['lat']}, lon: {val['lon']}, alt: {val['alt']}m)")
    print("[ 0] Unos proizvoljnih koordinata (bilo koja druga lokacija)")
    
    while True:
        try:
            izbor = int(input("\nIzaberite broj stanice (0-30): "))
            if izbor in STANICE:
                st = STANICE[izbor]
                break
            elif izbor == 0:
                naziv = input("Unesite naziv lokacije: ").strip().upper()
                lat = float(input("Unesite geografsku širinu (npr. 44.81): "))
                lon = float(input("Unesite geografsku dužinu (npr. 20.46): "))
                alt = int(input("Unesite nadmorsku visinu u metrima (npr. 120): "))
                st = {"naziv": naziv, "lat": lat, "lon": lon, "alt": alt}
                break
            else:
                print("Nevažeći izbor. Pokušajte ponovo.")
        except ValueError:
            print("Pogrešan unos! Unesite broj.")

    print("\n--- Unos vremenskog perioda (Dostupno od 1940. godine) ---")
    while True:
        try:
            start_year = int(input("Početna godina (npr. 1961 ili 1991): "))
            end_year = int(input("Završna godina (npr. 2020 ili 2025): "))
            if 1940 <= start_year <= end_year <= 2025:
                break
            else:
                print("Godine moraju biti u opsegu od 1940 do 2025!")
        except ValueError:
            print("Pogrešan unos! Unesite godinu kao četvorocifren broj.")
            
    return st, start_year, end_year

# ==============================================================================
# PREUZIMANJE I OBRADA PODATAKA (OPEN-METEO ERA5)
# ==============================================================================
st, START_YEAR, END_YEAR = izaberi_stanicu_i_period()
LOKACIJA_NAZIV = st["naziv"]
LATITUDE = st["lat"]
LONGITUDE = st["lon"]
ELEVATION_M = st["alt"]

print(f"\nPreuzimanje podataka sa Open-Meteo servera za {LOKACIJA_NAZIV} ({START_YEAR}-{END_YEAR})... Molimo sačekajte.")

url = "https://archive-api.open-meteo.com/v1/archive"
params = {
    "latitude": LATITUDE,
    "longitude": LONGITUDE,
    "start_date": f"{START_YEAR}-01-01",
    "end_date": f"{END_YEAR}-12-31",
    "daily": [
        "temperature_2m_mean", "temperature_2m_max", "temperature_2m_min",
        "relative_humidity_2m_mean", "sunshine_duration", "cloud_cover_mean",
        "precipitation_sum", "weather_code", "snowfall_sum"
    ],
    "timezone": "Europe/Belgrade"
}

resp = requests.get(url, params=params)
if resp.status_code != 200:
    raise Exception(f"Greška pri preuzimanju podataka sa mreže: {resp.status_code}")

df = pd.DataFrame(resp.json()["daily"])
df["time"] = pd.to_datetime(df["time"])
df["year"] = df["time"].dt.year
df["month"] = df["time"].dt.month

# Izračunavanje pomoćnih veličina
df["sunshine_hours"] = df["sunshine_duration"] / 3600.0
df["is_frost"] = df["temperature_2m_min"] < 0.0
df["is_tropical"] = df["temperature_2m_max"] >= 30.0
df["is_clear"] = df["cloud_cover_mean"] < 20.0
df["is_cloudy"] = df["cloud_cover_mean"] > 80.0
df["is_precip_01"] = df["precipitation_sum"] >= 0.1
df["is_precip_10"] = df["precipitation_sum"] >= 10.0
df["is_snow"] = df["weather_code"].isin([71, 73, 75, 77, 85, 86]) | (df["snowfall_sum"] > 0)
df["is_snow_cover"] = df["snowfall_sum"] > 0
df["is_fog"] = df["weather_code"].isin([45, 48])
df["is_hail"] = df["weather_code"].isin([89, 90, 96, 99])

months_names = ["jan", "feb", "mar", "apr", "maj", "jun", "jul", "avg", "sep", "okt", "nov", "dec"]

def calculate_yearly_table(df_year):
    metrics = {k: [] for k in [
        "norm_temp", "sred_max", "sred_min", "aps_max", "aps_min", "mrazni_dani", "tropski_dani",
        "rel_vlaga", "suns_hrs", "vedri_dani", "oblacni_dani", "padavine_suma", "max_dnev_pad",
        "pad_01mm", "pad_10mm", "sneg_dani", "sneg_pokrivac", "magla_dani", "grad_dani"
    ]}
    
    for m in range(1, 13):
        m_df = df_year[df_year["month"] == m]
        if len(m_df) == 0:
            for k in metrics: metrics[k].append(0)
            continue
            
        metrics["norm_temp"].append(round(m_df["temperature_2m_mean"].mean(), 1))
        metrics["sred_max"].append(round(m_df["temperature_2m_max"].mean(), 1))
        metrics["sred_min"].append(round(m_df["temperature_2m_min"].mean(), 1))
        metrics["aps_max"].append(round(m_df["temperature_2m_max"].max(), 1))
        metrics["aps_min"].append(round(m_df["temperature_2m_min"].min(), 1))
        metrics["mrazni_dani"].append(int(m_df["is_frost"].sum()))
        metrics["tropski_dani"].append(int(m_df["is_tropical"].sum()))
        metrics["rel_vlaga"].append(round(m_df["relative_humidity_2m_mean"].mean(), 1))
        metrics["suns_hrs"].append(round(m_df["sunshine_hours"].sum(), 1))
        metrics["vedri_dani"].append(int(m_df["is_clear"].sum()))
        metrics["oblacni_dani"].append(int(m_df["is_cloudy"].sum()))
        metrics["padavine_suma"].append(round(m_df["precipitation_sum"].sum(), 1))
        metrics["max_dnev_pad"].append(round(m_df["precipitation_sum"].max(), 1))
        metrics["pad_01mm"].append(int(m_df["is_precip_01"].sum()))
        metrics["pad_10mm"].append(int(m_df["is_precip_10"].sum()))
        metrics["sneg_dani"].append(int(m_df["is_snow"].sum()))
        metrics["sneg_pokrivac"].append(int(m_df["is_snow_cover"].sum()))
        metrics["magla_dani"].append(int(m_df["is_fog"].sum()))
        metrics["grad_dani"].append(int(m_df["is_hail"].sum()))
        
    return metrics

years_list = list(range(START_YEAR, END_YEAR + 1))
yearly_results = {y: calculate_yearly_table(df[df["year"] == y]) for y in years_list}

# ==============================================================================
# GENERISANJE EXCEL FAJLA MEĐUSOBNO POVEZANOG FORMULAMA
# ==============================================================================
wb = openpyxl.Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
SECTION_FILL = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
TITLE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
                     top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
HEADER_BORDER = Border(left=Side(style='thin', color='808080'), right=Side(style='thin', color='808080'),
                       top=Side(style='thin', color='808080'), bottom=Side(style='medium', color='000000'))

rows_schema = [
    (5, 'section', 'TEMPERATURA VAZDUHA (°C)', None),
    (6, 'metric', 'Normalna vrednost', 'AVG'),
    (7, 'metric', 'Srednja maksimalna', 'AVG'),
    (8, 'metric', 'Srednja minimalna', 'AVG'),
    (9, 'metric', 'Apsolutni maksimum', 'MAX'),
    (10, 'metric', 'Apsolutni minimum', 'MIN'),
    (11, 'metric', 'Sr. br. mraznih dana', 'SUM'),
    (12, 'metric', 'Sr. br. tropskih dana', 'SUM'),
    (13, 'section', 'RELATIVNA VLAGA (%)', None),
    (14, 'metric', 'Prosek', 'AVG'),
    (15, 'section', 'TRAJANJE SIJANJA SUNCA (h)', None),
    (16, 'metric', 'Prosek', 'SUM'),
    (17, 'metric', 'Broj vedrih dana', 'SUM'),
    (18, 'metric', 'Broj oblačnih dana', 'SUM'),
    (19, 'section', 'PADAVINE (mm)', None),
    (20, 'metric', 'Sr. mesečna suma', 'SUM'),
    (21, 'metric', 'Max. dnevna suma', 'MAX'),
    (22, 'metric', 'Sr. br. dana >= 0.1 mm', 'SUM'),
    (23, 'metric', 'Sr. br. dana >= 10.0 mm', 'SUM'),
    (24, 'section', 'POJAVE (broj dana sa....)', None),
    (25, 'metric', 'snegom', 'SUM'),
    (26, 'metric', 'snežnim pokrivačem', 'SUM'),
    (27, 'metric', 'maglom', 'SUM'),
    (28, 'metric', 'gradom', 'SUM')
]

metric_keys = [
    "norm_temp", "sred_max", "sred_min", "aps_max", "aps_min", "mrazni_dani", "tropski_dani",
    "rel_vlaga", "suns_hrs", "vedri_dani", "oblacni_dani",
    "padavine_suma", "max_dnev_pad", "pad_01mm", "pad_10mm",
    "sneg_dani", "sneg_pokrivac", "magla_dani", "grad_dani"
]

def build_excel_sheet(ws, title_text, subtitle_text, data_dict=None, is_summary=False):
    ws.views.sheetView[0].showGridLines = True
    
    ws.merge_cells("A1:N1")
    ws["A1"] = title_text
    ws["A1"].font = Font(name="Calibri", size=11, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = TITLE_FILL
    
    ws.merge_cells("A2:N2")
    ws["A2"] = subtitle_text
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws["A4"] = "METEOROLOŠKI PARAMETAR"
    ws["A4"].font = Font(name="Calibri", size=10, bold=True)
    ws["A4"].fill = HEADER_FILL
    ws["A4"].border = HEADER_BORDER
    
    cols = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    for i, m_name in enumerate(months_names):
        col_let = cols[i]
        c_cell = ws[f"{col_let}4"]
        c_cell.value = m_name
        c_cell.font = Font(name="Calibri", size=10, bold=True)
        c_cell.fill = HEADER_FILL
        c_cell.alignment = Alignment(horizontal="center", vertical="center")
        c_cell.border = HEADER_BORDER
        
    ws["N4"] = "god."
    ws["N4"].font = Font(name="Calibri", size=10, bold=True)
    ws["N4"].fill = HEADER_FILL
    ws["N4"].alignment = Alignment(horizontal="center", vertical="center")
    ws["N4"].border = HEADER_BORDER
    
    m_idx = 0
    for r_idx, r_type, label, agg_type in rows_schema:
        if r_type == 'section':
            ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=14)
            cell = ws.cell(row=r_idx, column=1, value=label)
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.fill = SECTION_FILL
            for c in range(1, 15):
                ws.cell(row=r_idx, column=c).border = THIN_BORDER
        else:
            cell = ws.cell(row=r_idx, column=1, value=label)
            cell.font = Font(name="Calibri", size=10)
            cell.border = THIN_BORDER
            
            m_key = metric_keys[m_idx]
            m_idx += 1
            
            for c_i, col_let in enumerate(cols):
                c_cell = ws[f"{col_let}{r_idx}"]
                c_cell.border = THIN_BORDER
                c_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                if is_summary:
                    refs = [f"'{y}'!{col_let}{r_idx}" for y in years_list]
                    c_cell.value = f"=AVERAGE({','.join(refs)})"
                else:
                    c_cell.value = data_dict[m_key][c_i]
                
                if m_key in ["mrazni_dani", "tropski_dani", "vedri_dani", "oblacni_dani", 
                             "pad_01mm", "pad_10mm", "sneg_dani", "sneg_pokrivac", "magla_dani", "grad_dani"]:
                    c_cell.number_format = "0.0" if is_summary else "0"
                else:
                    c_cell.number_format = "0.0"
            
            n_cell = ws[f"N{r_idx}"]
            n_cell.border = THIN_BORDER
            n_cell.alignment = Alignment(horizontal="center", vertical="center")
            n_cell.font = Font(name="Calibri", size=10, bold=True)
            
            if agg_type == 'AVG': n_cell.value = f"=AVERAGE(B{r_idx}:M{r_idx})"
            elif agg_type == 'SUM': n_cell.value = f"=SUM(B{r_idx}:M{r_idx})"
            elif agg_type == 'MAX': n_cell.value = f"=MAX(B{r_idx}:M{r_idx})"
            elif agg_type == 'MIN': n_cell.value = f"=MIN(B{r_idx}:M{r_idx})"
            
            if m_key in ["mrazni_dani", "tropski_dani", "vedri_dani", "oblacni_dani", 
                         "pad_01mm", "pad_10mm", "sneg_dani", "sneg_pokrivac", "magla_dani", "grad_dani"]:
                n_cell.number_format = "0.0" if is_summary else "0"
            else:
                n_cell.number_format = "0.0"

    ws.column_dimensions['A'].width = 30
    for col_let in cols + ["N"]:
        ws.column_dimensions[col_let].width = 9

# 1. Sheet sa višegodišnjim prosekom
ws_sum = wb.create_sheet(title=f"Prosek {START_YEAR}-{END_YEAR}")
build_excel_sheet(
    ws_sum,
    title_text=f"PROSEČNE MESEČNE, GODIŠNJE I EKSTREMNE VREDNOSTI ZA PERIOD {START_YEAR}-{END_YEAR}. GODINA",
    subtitle_text=f"{LOKACIJA_NAZIV}  φ {LATITUDE}°N  λ {LONGITUDE}°E  h {ELEVATION_M} m",
    is_summary=True
)

# 2. Pojedinačni sheet-ovi po godinama
for y in years_list:
    ws_y = wb.create_sheet(title=str(y))
    build_excel_sheet(
        ws_y,
        title_text=f"PROSEČNE MESEČNE, GODIŠNJE I EKSTREMNE VREDNOSTI ZA {y}. GODINU",
        subtitle_text=f"{LOKACIJA_NAZIV}  φ {LATITUDE}°N  λ {LONGITUDE}°E  h {ELEVATION_M} m",
        data_dict=yearly_results[y],
        is_summary=False
    )

safe_filename = LOKACIJA_NAZIV.replace(" ", "_").replace("-", "_")
output_filename = f"Klimatoloski_Podaci_{safe_filename}_{START_YEAR}_{END_YEAR}.xlsx"
wb.save(output_filename)

print(f"\n[USPEH] Generisan je kompletan Excel fajl: {output_filename}")