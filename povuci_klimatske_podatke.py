import requests
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==========================================
# 1. PODEŠAVANJE LOKACIJE I PERIODA
# ==========================================
LOKACIJA_NAZIV = "SJENICA"
LATITUDE = 43.27
LONGITUDE = 19.99
ELEVATION_M = 1038

START_YEAR = 2021
END_YEAR = 2025

OUTPUT_EXCEL = f"Klimatoloski_Podaci_{LOKACIJA_NAZIV}_{START_YEAR}_{END_YEAR}.xlsx"
OUTPUT_CSV = f"Klimatoloski_Podaci_{LOKACIJA_NAZIV}_{START_YEAR}_{END_YEAR}.csv"

# ==========================================
# 2. PREUZIMANJE DNEVNIH PODATAKA (API)
# ==========================================
print(f"Preuzimanje klimatoloških podataka za {LOKACIJA_NAZIV} ({START_YEAR}-{END_YEAR})...")

url = "https://archive-api.open-meteo.com/v1/archive"
params = {
    "latitude": LATITUDE,
    "longitude": LONGITUDE,
    "start_date": f"{START_YEAR}-01-01",
    "end_date": f"{END_YEAR}-12-31",
    "daily": [
        "temperature_2m_mean",
        "temperature_2m_max",
        "temperature_2m_min",
        "relative_humidity_2m_mean",
        "sunshine_duration",
        "cloud_cover_mean",
        "precipitation_sum",
        "weather_code",
        "snowfall_sum"
    ],
    "timezone": "Europe/Belgrade"
}

response = requests.get(url, params=params)
if response.status_code != 200:
    raise Exception(f"Greška pri preuzimanju podataka: {response.status_code}")

data = response.json()["daily"]
df = pd.DataFrame(data)
df["time"] = pd.to_datetime(df["time"])
df["year"] = df["time"].dt.year
df["month"] = df["time"].dt.month

# Prevaranje osunčavanja iz sekundi u sate
df["sunshine_hours"] = df["sunshine_duration"] / 3600.0

# Indikatorske kolone za pojave
df["is_frost"] = df["temperature_2m_min"] < 0.0
df["is_tropical"] = df["temperature_2m_max"] >= 30.0
df["is_clear"] = df["cloud_cover_mean"] < 20.0
df["is_cloudy"] = df["cloud_cover_mean"] > 80.0
df["is_precip_01"] = df["precipitation_sum"] >= 0.1
df["is_precip_10"] = df["precipitation_sum"] >= 10.0

# WMO weather codes za sneg, maglu i grad
df["is_snow"] = df["weather_code"].isin([71, 73, 75, 77, 85, 86]) | (df["snowfall_sum"] > 0)
df["is_snow_cover"] = df["snowfall_sum"] > 0
df["is_fog"] = df["weather_code"].isin([45, 48])
df["is_hail"] = df["weather_code"].isin([89, 90, 96, 99])

# ==========================================
# 3. OBRAĐIVANJE MESECNIH PARAMETARA
# ==========================================
months_names = ["jan", "feb", "mar", "apr", "maj", "jun", "jul", "avg", "sep", "okt", "nov", "dec"]

def calculate_yearly_table(df_year):
    table_data = {}
    
    # Inicijalizacija lista po mesecima
    metrics = {
        "norm_temp": [], "sred_max": [], "sred_min": [], "aps_max": [], "aps_min": [],
        "mrazni_dani": [], "tropski_dani": [], "rel_vlaga": [], "suns_hrs": [],
        "vedri_dani": [], "oblacni_dani": [], "padavine_suma": [], "max_dnev_pad": [],
        "pad_01mm": [], "pad_10mm": [], "sneg_dani": [], "sneg_pokrivac": [],
        "magla_dani": [], "grad_dani": []
    }
    
    for m in range(1, 13):
        m_df = df_year[df_year["month"] == m]
        if len(m_df) == 0:
            for k in metrics: metrics[k].append(0)
            continue
            
        metrics["norm_temp"].append(round(m_df["temperature_2m_mean"].mean(), 1))
        metrics["sred_max"].append(round(m_df["temperature_2m_max"].mean(), 1))
        metrics["sred_min"].append(round(m_df["temperature_2m_min"].mean(), 1))
        metrics["aps_max"].append(round(m_df["temperature_2m_max"].max(), 1))
        metrics["aps_min"].append(round(m_df["temperature_2m_min"].min(), 1))
        metrics["mrazni_dani"].append(int(m_df["is_frost"].sum()))
        metrics["tropski_dani"].append(int(m_df["is_tropical"].sum()))
        
        metrics["rel_vlaga"].append(round(m_df["relative_humidity_2m_mean"].mean(), 1))
        
        metrics["suns_hrs"].append(round(m_df["sunshine_hours"].sum(), 1))
        metrics["vedri_dani"].append(int(m_df["is_clear"].sum()))
        metrics["oblacni_dani"].append(int(m_df["is_cloudy"].sum()))
        
        metrics["padavine_suma"].append(round(m_df["precipitation_sum"].sum(), 1))
        metrics["max_dnev_pad"].append(round(m_df["precipitation_sum"].max(), 1))
        metrics["pad_01mm"].append(int(m_df["is_precip_01"].sum()))
        metrics["pad_10mm"].append(int(m_df["is_precip_10"].sum()))
        
        metrics["sneg_dani"].append(int(m_df["is_snow"].sum()))
        metrics["sneg_pokrivac"].append(int(m_df["is_snow_cover"].sum()))
        metrics["magla_dani"].append(int(m_df["is_fog"].sum()))
        metrics["grad_dani"].append(int(m_df["is_hail"].sum()))
        
    return metrics

yearly_results = {}
for y in range(START_YEAR, END_YEAR + 1):
    df_y = df[df["year"] == y]
    yearly_results[y] = calculate_yearly_table(df_y)

# ==========================================
# 4. FORMIRANJE LEPO FORMATIRANOG EXCEL-A
# ==========================================
wb = openpyxl.Workbook()
wb.remove(wb.active) # Uklanjanje podrazumevanog sheet-a

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
SECTION_FILL = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
TITLE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
)
HEADER_BORDER = Border(
    left=Side(style='thin', color='808080'), right=Side(style='thin', color='808080'),
    top=Side(style='thin', color='808080'), bottom=Side(style='medium', color='000000')
)

rows_schema = [
    (5, 'section', 'TEMPERATURA VAZDUHA (°C)', None),
    (6, 'metric', 'Normalna vrednost', 'AVG'),
    (7, 'metric', 'Srednja maksimalna', 'AVG'),
    (8, 'metric', 'Srednja minimalna', 'AVG'),
    (9, 'metric', 'Apsolutni maksimum', 'MAX'),
    (10, 'metric', 'Apsolutni minimum', 'MIN'),
    (11, 'metric', 'Sr. br. mraznih dana', 'SUM'),
    (12, 'metric', 'Sr. br. tropskih dana', 'SUM'),
    (13, 'section', 'RELATIVNA VLAGA (%)', None),
    (14, 'metric', 'Prosek', 'AVG'),
    (15, 'section', 'TRAJANJE SIJANJA SUNCA (h)', None),
    (16, 'metric', 'Prosek', 'SUM'),
    (17, 'metric', 'Broj vedrih dana', 'SUM'),
    (18, 'metric', 'Broj oblačnih dana', 'SUM'),
    (19, 'section', 'PADAVINE (mm)', None),
    (20, 'metric', 'Sr. mesečna suma', 'SUM'),
    (21, 'metric', 'Max. dnevna suma', 'MAX'),
    (22, 'metric', 'Sr. br. dana >= 0.1 mm', 'SUM'),
    (23, 'metric', 'Sr. br. dana >= 10.0 mm', 'SUM'),
    (24, 'section', 'POJAVE (broj dana sa....)', None),
    (25, 'metric', 'snegom', 'SUM'),
    (26, 'metric', 'snežnim pokrivačem', 'SUM'),
    (27, 'metric', 'maglom', 'SUM'),
    (28, 'metric', 'gradom', 'SUM')
]

metric_keys = [
    "norm_temp", "sred_max", "sred_min", "aps_max", "aps_min", "mrazni_dani", "tropski_dani",
    "rel_vlaga", "suns_hrs", "vedri_dani", "oblacni_dani",
    "padavine_suma", "max_dnev_pad", "pad_01mm", "pad_10mm",
    "sneg_dani", "sneg_pokrivac", "magla_dani", "grad_dani"
]

def build_excel_sheet(ws, title_text, subtitle_text, data_dict=None, is_summary=False, year_list=None):
    ws.views.sheetView[0].showGridLines = True
    
    ws.merge_cells("A1:N1")
    ws["A1"] = title_text
    ws["A1"].font = Font(name="Calibri", size=11, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = TITLE_FILL
    
    ws.merge_cells("A2:N2")
    ws["A2"] = subtitle_text
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws["A4"] = "METEOROLOŠKI PARAMETAR"
    ws["A4"].font = Font(name="Calibri", size=10, bold=True)
    ws["A4"].fill = HEADER_FILL
    ws["A4"].border = HEADER_BORDER
    
    cols = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    for i, m_name in enumerate(months_names):
        col_let = cols[i]
        c_cell = ws[f"{col_let}4"]
        c_cell.value = m_name
        c_cell.font = Font(name="Calibri", size=10, bold=True)
        c_cell.fill = HEADER_FILL
        c_cell.alignment = Alignment(horizontal="center", vertical="center")
        c_cell.border = HEADER_BORDER
        
    ws["N4"] = "god."
    ws["N4"].font = Font(name="Calibri", size=10, bold=True)
    ws["N4"].fill = HEADER_FILL
    ws["N4"].alignment = Alignment(horizontal="center", vertical="center")
    ws["N4"].border = HEADER_BORDER
    
    m_idx = 0
    for r_idx, r_type, label, agg_type in rows_schema:
        if r_type == 'section':
            ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=14)
            cell = ws.cell(row=r_idx, column=1, value=label)
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.fill = SECTION_FILL
            for c in range(1, 15):
                ws.cell(row=r_idx, column=c).border = THIN_BORDER
        else:
            cell = ws.cell(row=r_idx, column=1, value=label)
            cell.font = Font(name="Calibri", size=10)
            cell.border = THIN_BORDER
            
            m_key = metric_keys[m_idx]
            m_idx += 1
            
            for c_i, col_let in enumerate(cols):
                c_cell = ws[f"{col_let}{r_idx}"]
                c_cell.border = THIN_BORDER
                c_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                if is_summary:
                    refs = [f"'{y}'!{col_let}{r_idx}" for y in year_list]
                    c_cell.value = f"=AVERAGE({','.join(refs)})"
                else:
                    c_cell.value = data_dict[m_key][c_i]
                
                if m_key in ["mrazni_dani", "tropski_dani", "vedri_dani", "oblacni_dani", 
                             "pad_01mm", "pad_10mm", "sneg_dani", "sneg_pokrivac", "magla_dani", "grad_dani"]:
                    c_cell.number_format = "0.0" if is_summary else "0"
                else:
                    c_cell.number_format = "0.0"
            
            n_cell = ws[f"N{r_idx}"]
            n_cell.border = THIN_BORDER
            n_cell.alignment = Alignment(horizontal="center", vertical="center")
            n_cell.font = Font(name="Calibri", size=10, bold=True)
            
            if agg_type == 'AVG': n_cell.value = f"=AVERAGE(B{r_idx}:M{r_idx})"
            elif agg_type == 'SUM': n_cell.value = f"=SUM(B{r_idx}:M{r_idx})"
            elif agg_type == 'MAX': n_cell.value = f"=MAX(B{r_idx}:M{r_idx})"
            elif agg_type == 'MIN': n_cell.value = f"=MIN(B{r_idx}:M{r_idx})"
            
            if m_key in ["mrazni_dani", "tropski_dani", "vedri_dani", "oblacni_dani", 
                         "pad_01mm", "pad_10mm", "sneg_dani", "sneg_pokrivac", "magla_dani", "grad_dani"]:
                n_cell.number_format = "0.0" if is_summary else "0"
            else:
                n_cell.number_format = "0.0"

    ws.column_dimensions['A'].width = 30
    for col_let in cols + ["N"]:
        ws.column_dimensions[col_let].width = 9

# Pravljenje Zbirnog Sheet-a
years_list = list(range(START_YEAR, END_YEAR + 1))
ws_sum = wb.create_sheet(title=f"Prosek {START_YEAR}-{END_YEAR}")
build_excel_sheet(
    ws_sum,
    title_text=f"PROSEČNE MESEČNE, GODIŠNJE I EKSTREMNE VREDNOSTI ZA PERIOD {START_YEAR}-{END_YEAR}. GODINA",
    subtitle_text=f"{LOKACIJA_NAZIV}  φ {LATITUDE}°N  λ {LONGITUDE}°E  h {ELEVATION_M} m",
    is_summary=True,
    year_list=years_list
)

# Pravljenje Pojedinačnih Sheet-ova
for y in years_list:
    ws_y = wb.create_sheet(title=str(y))
    build_excel_sheet(
        ws_y,
        title_text=f"PROSEČNE MESEČNE, GODIŠNJE I EKSTREMNE VREDNOSTI ZA {y}. GODINU",
        subtitle_text=f"{LOKACIJA_NAZIV}  φ {LATITUDE}°N  λ {LONGITUDE}°E  h {ELEVATION_M} m",
        data_dict=yearly_results[y],
        is_summary=False
    )

wb.save(OUTPUT_EXCEL)
print(f"Uspešno generisan Excel fajl: {OUTPUT_EXCEL}")